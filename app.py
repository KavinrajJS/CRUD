from flask import Flask, render_template_string, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = "employee_management_secret"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "sheets", "Employees.xlsx")


# -----------------------------
# Create Excel File if Missing
# -----------------------------
def create_excel_file():
    sheets_folder = os.path.join(BASE_DIR, "sheets")
    if not os.path.exists(sheets_folder):
        os.makedirs(sheets_folder)

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Employee Name", "Department", "Salary"])
        wb.save(EXCEL_FILE)
        wb.close()
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(["Employee ID", "Employee Name", "Department", "Salary"])
            wb.save(EXCEL_FILE)
        wb.close()


# -----------------------------
# Get All Employees
# -----------------------------
def get_all_employees():
    create_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            employees.append({
                "id": row[0],
                "name": row[1],
                "department": row[2],
                "salary": row[3]
            })
    wb.close()
    return employees


# -----------------------------
# Dashboard Statistics
# -----------------------------
def get_dashboard_stats():
    employees = get_all_employees()
    total_employees = len(employees)
    departments = set()
    total_salary = 0

    for emp in employees:
        if emp["department"]:
            departments.add(emp["department"])
        try:
            total_salary += float(emp["salary"])
        except:
            pass

    avg_salary = round(total_salary / total_employees, 2) if total_employees > 0 else 0
    return {
        "total_employees": total_employees,
        "total_departments": len(departments),
        "avg_salary": avg_salary
    }


# ==========================================
# HTML + CSS Templates (Consolidated)
# ==========================================

INDEX_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Employee Management System</title>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0f172a;
            --bg-gradient: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%);
            --card-bg: rgba(30, 41, 59, 0.45);
            --glass-border: rgba(255, 255, 255, 0.08);
            --glass-border-focus: rgba(56, 189, 248, 0.4);
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --accent: #38bdf8;
            --accent-hover: #0ea5e9;
            --accent-gradient: linear-gradient(135deg, #38bdf8 0%, #2563eb 100%);
            --success: #10b981;
            --danger: #ef4444;
        }

        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: 'Poppins', sans-serif;
        }

        body {
            background: var(--bg-gradient);
            background-attachment: fixed;
            min-height: 100vh;
            color: var(--text-primary);
            overflow-x: hidden;
        }

        /* Header */
        .header {
            width: 100%;
            padding: 20px 40px;
            background: rgba(15, 23, 42, 0.6);
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
            border-bottom: 1px solid var(--glass-border);
            position: sticky;
            top: 0;
            z-index: 1000;
        }

        .logo {
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 22px;
            font-weight: 700;
            letter-spacing: 0.5px;
        }

        .logo i {
            color: var(--accent);
            filter: drop-shadow(0 0 8px rgba(56, 189, 248, 0.5));
        }

        /* Layout Container */
        .container {
            width: 90%;
            max-width: 1400px;
            margin: 40px auto;
        }

        /* Dashboard Cards */
        .dashboard {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 25px;
            margin-bottom: 40px;
            animation: fadeInUp 0.6s ease-out;
        }

        .card {
            background: var(--card-bg);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }

        .card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 4px;
            background: var(--accent-gradient);
            opacity: 0;
            transition: opacity 0.3s ease;
        }

        .card:hover::before {
            opacity: 1;
        }

        .card:hover {
            transform: translateY(-5px);
            border-color: var(--glass-border-focus);
            box-shadow: 0 20px 40px rgba(56, 189, 248, 0.1);
        }

        .card h3 {
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: var(--text-secondary);
            margin-bottom: 12px;
        }

        .card h1 {
            font-size: 32px;
            font-weight: 700;
            color: var(--text-primary);
        }

        /* Split Screen Grid */
        .content-grid {
            display: grid;
            grid-template-columns: 360px 1fr;
            gap: 30px;
            align-items: start;
            animation: fadeInUp 0.8s ease-out;
        }

        @media (max-width: 1024px) {
            .content-grid {
                grid-template-columns: 1fr;
            }
        }

        /* Form Card */
        .form-card {
            background: var(--card-bg);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
        }

        .form-card h2 {
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 24px;
            color: var(--text-primary);
            border-left: 4px solid var(--accent);
            padding-left: 12px;
        }

        .form-group {
            margin-bottom: 20px;
        }

        .form-group label {
            display: block;
            font-size: 13px;
            font-weight: 500;
            color: var(--text-secondary);
            margin-bottom: 8px;
        }

        .form-group input {
            width: 100%;
            padding: 12px 16px;
            background: rgba(15, 23, 42, 0.4);
            border: 1px solid var(--glass-border);
            border-radius: 10px;
            color: var(--text-primary);
            outline: none;
            font-size: 14px;
            transition: all 0.3s ease;
        }

        .form-group input::placeholder {
            color: #64748b;
        }

        .form-group input:focus {
            border-color: var(--accent);
            background: rgba(15, 23, 42, 0.6);
            box-shadow: 0 0 0 3px rgba(56, 189, 248, 0.15);
        }

        .btn-primary {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            width: 100%;
            background: var(--accent-gradient);
            color: white;
            border: none;
            padding: 14px;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s ease;
        }

        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(37, 99, 235, 0.3);
        }

        /* Table Card */
        .table-card {
            background: var(--card-bg);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
        }

        .table-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 15px;
            margin-bottom: 24px;
        }

        .table-header h2 {
            font-size: 20px;
            font-weight: 600;
            color: var(--text-primary);
            border-left: 4px solid var(--accent);
            padding-left: 12px;
        }

        .search-wrapper {
            position: relative;
            width: 280px;
        }

        @media (max-width: 640px) {
            .search-wrapper {
                width: 100%;
            }
        }

        .search-wrapper i {
            position: absolute;
            left: 14px;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-secondary);
            font-size: 14px;
        }

        .search-wrapper input {
            width: 100%;
            padding: 10px 16px 10px 38px;
            background: rgba(15, 23, 42, 0.4);
            border: 1px solid var(--glass-border);
            border-radius: 10px;
            color: var(--text-primary);
            outline: none;
            font-size: 14px;
            transition: all 0.3s ease;
        }

        .search-wrapper input:focus {
            border-color: var(--accent);
            background: rgba(15, 23, 42, 0.6);
            box-shadow: 0 0 0 3px rgba(56, 189, 248, 0.15);
        }

        .table-wrapper {
            overflow-x: auto;
            border-radius: 12px;
            border: 1px solid var(--glass-border);
        }

        table {
            width: 100%;
            border-collapse: collapse;
            text-align: left;
        }

        thead {
            background: rgba(15, 23, 42, 0.5);
            border-bottom: 2px solid var(--glass-border);
        }

        th {
            padding: 16px;
            font-size: 13px;
            font-weight: 600;
            color: var(--text-secondary);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        td {
            padding: 16px;
            font-size: 14px;
            color: var(--text-primary);
            border-bottom: 1px solid var(--glass-border);
        }

        tbody tr {
            transition: all 0.2s ease;
        }

        tbody tr:hover {
            background: rgba(255, 255, 255, 0.03);
        }

        tbody tr:last-child td {
            border-bottom: none;
        }

        /* Actions */
        .btn-edit, .btn-delete {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            text-decoration: none;
            padding: 6px 12px;
            border-radius: 6px;
            font-size: 13px;
            font-weight: 500;
            transition: all 0.2s ease;
        }

        .btn-edit {
            background: rgba(56, 189, 248, 0.15);
            color: var(--accent);
            margin-right: 8px;
        }

        .btn-edit:hover {
            background: var(--accent);
            color: #0f172a;
        }

        .btn-delete {
            background: rgba(239, 68, 68, 0.15);
            color: var(--danger);
        }

        .btn-delete:hover {
            background: var(--danger);
            color: white;
        }

        /* Alerts */
        .alert-container {
            margin-bottom: 25px;
            animation: fadeIn 0.5s ease-out;
        }

        .alert {
            padding: 14px 20px;
            border-radius: 10px;
            font-weight: 500;
            font-size: 14px;
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 10px;
            border: 1px solid transparent;
        }

        .alert.success {
            background: rgba(16, 185, 129, 0.15);
            color: var(--success);
            border-color: rgba(16, 185, 129, 0.2);
        }

        .alert.danger {
            background: rgba(239, 68, 68, 0.15);
            color: var(--danger);
            border-color: rgba(239, 68, 68, 0.2);
        }

        /* Animations */
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
    </style>
</head>
<body>

    <!-- Header -->
    <header class="header">
        <div class="logo">
            <i class="fas fa-users"></i>
            <span>Employee Management System</span>
        </div>
    </header>

    <div class="container">

        <!-- Flash Messages -->
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                <div class="alert-container">
                    {% for category, message in messages %}
                        <div class="alert {{ category }}">
                            {% if category == 'success' %}
                                <i class="fas fa-check-circle"></i>
                            {% else %}
                                <i class="fas fa-exclamation-circle"></i>
                            {% endif %}
                            {{ message }}
                        </div>
                    {% endfor %}
                </div>
            {% endif %}
        {% endwith %}

        <!-- Dashboard Cards -->
        <section class="dashboard">
            <div class="card">
                <h3>Total Employees</h3>
                <h1>{{ stats.total_employees }}</h1>
            </div>

            <div class="card">
                <h3>Total Departments</h3>
                <h1>{{ stats.total_departments }}</h1>
            </div>

            <div class="card">
                <h3>Average Salary</h3>
                <h1>₹{{ stats.avg_salary }}</h1>
            </div>
        </section>

        <!-- Main Workspace: Form and Table Split -->
        <div class="content-grid">

            <!-- Add Employee Form -->
            <section class="form-section">
                <div class="form-card">
                    <h2>Add Employee</h2>
                    <form action="/add" method="POST">
                        <div class="form-group">
                            <label>Employee ID</label>
                            <input
                                type="number"
                                name="employee_id"
                                min="1"
                                placeholder="Enter Employee ID"
                                required>
                        </div>

                        <div class="form-group">
                            <label>Employee Name</label>
                            <input
                                type="text"
                                name="employee_name"
                                placeholder="Enter Employee Name"
                                required>
                        </div>

                        <div class="form-group">
                            <label>Department</label>
                            <input
                                type="text"
                                name="department"
                                placeholder="Enter Department"
                                required>
                        </div>

                        <div class="form-group">
                            <label>Salary</label>
                            <input
                                type="number"
                                name="salary"
                                placeholder="Enter Salary"
                                required>
                        </div>

                        <button type="submit" class="btn-primary">
                            <i class="fas fa-plus"></i>
                            Add Employee
                        </button>
                    </form>
                </div>
            </section>

            <!-- Employee Table -->
            <section class="table-section">
                <div class="table-card">
                    <div class="table-header">
                        <h2>Employee Records</h2>
                        <div class="search-wrapper">
                            <i class="fas fa-search"></i>
                            <input
                                type="text"
                                id="searchInput"
                                placeholder="Search employees..."
                                onkeyup="searchEmployee()">
                        </div>
                    </div>

                    <div class="table-wrapper">
                        <table id="employeeTable">
                            <thead>
                                <tr>
                                    <th>ID</th>
                                    <th>Name</th>
                                    <th>Department</th>
                                    <th>Salary</th>
                                    <th>Actions</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for emp in employees %}
                                <tr>
                                    <td>{{ emp.id }}</td>
                                    <td>{{ emp.name }}</td>
                                    <td>{{ emp.department }}</td>
                                    <td>₹{{ emp.salary }}</td>
                                    <td>
                                        <a href="/edit/{{ emp.id }}" class="btn-edit">
                                            <i class="fas fa-edit"></i> Edit
                                        </a>
                                        <a
                                            href="/delete/{{ emp.id }}"
                                            class="btn-delete"
                                            onclick="return confirm('Are you sure you want to delete this employee?')">
                                            <i class="fas fa-trash"></i> Delete
                                        </a>
                                    </td>
                                </tr>
                                {% else %}
                                <tr>
                                    <td colspan="5" style="text-align: center; color: var(--text-secondary); padding: 30px;">
                                        No employee records found.
                                    </td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            </section>

        </div>
    </div>

    <script>
        function searchEmployee() {
            let input = document.getElementById("searchInput");
            let filter = input.value.toUpperCase();
            let table = document.getElementById("employeeTable");
            let tr = table.getElementsByTagName("tr");

            for (let i = 1; i < tr.length; i++) {
                // If the row is a "No records found" row, skip it
                if (tr[i].cells.length < 5) continue;
                
                let found = false;
                // Search across ID, Name, Department, and Salary columns
                for (let j = 0; j < 4; j++) {
                    let td = tr[i].getElementsByTagName("td")[j];
                    if (td) {
                        let txtValue = td.textContent || td.innerText;
                        if (txtValue.toUpperCase().indexOf(filter) > -1) {
                            found = true;
                            break;
                        }
                    }
                }
                tr[i].style.display = found ? "" : "none";
            }
        }
    </script>
</body>
</html>
"""

EDIT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Edit Employee</title>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0f172a;
            --bg-gradient: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%);
            --card-bg: rgba(30, 41, 59, 0.45);
            --glass-border: rgba(255, 255, 255, 0.08);
            --glass-border-focus: rgba(56, 189, 248, 0.4);
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --accent: #38bdf8;
            --accent-hover: #0ea5e9;
            --accent-gradient: linear-gradient(135deg, #38bdf8 0%, #2563eb 100%);
            --success: #10b981;
            --danger: #ef4444;
        }

        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: 'Poppins', sans-serif;
        }

        body {
            background: var(--bg-gradient);
            background-attachment: fixed;
            min-height: 100vh;
            color: var(--text-primary);
            overflow-x: hidden;
        }

        .edit-container {
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 30px;
            animation: fadeIn 0.5s ease-out;
        }

        .edit-card {
            width: 100%;
            max-width: 500px;
            background: var(--card-bg);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 40px;
            box-shadow: 0 20px 50px rgba(0, 0, 0, 0.3);
        }

        .edit-header {
            text-align: center;
            margin-bottom: 30px;
        }

        .edit-header i {
            font-size: 40px;
            color: var(--accent);
            margin-bottom: 12px;
            filter: drop-shadow(0 0 8px rgba(56, 189, 248, 0.4));
        }

        .edit-header h2 {
            font-size: 24px;
            font-weight: 600;
            color: var(--text-primary);
        }

        .form-group {
            margin-bottom: 20px;
        }

        .form-group label {
            display: block;
            font-size: 13px;
            font-weight: 500;
            color: var(--text-secondary);
            margin-bottom: 8px;
        }

        .form-group input {
            width: 100%;
            padding: 12px 16px;
            background: rgba(15, 23, 42, 0.4);
            border: 1px solid var(--glass-border);
            border-radius: 10px;
            color: var(--text-primary);
            outline: none;
            font-size: 14px;
            transition: all 0.3s ease;
        }

        .form-group input:focus {
            border-color: var(--accent);
            background: rgba(15, 23, 42, 0.6);
            box-shadow: 0 0 0 3px rgba(56, 189, 248, 0.15);
        }

        .readonly-input {
            background: rgba(15, 23, 42, 0.6) !important;
            cursor: not-allowed;
            color: var(--text-secondary) !important;
            border-color: var(--glass-border) !important;
        }

        .button-group {
            display: flex;
            gap: 15px;
            margin-top: 30px;
        }

        .btn-primary {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            flex: 1;
            background: var(--accent-gradient);
            color: white;
            border: none;
            padding: 14px;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s ease;
            text-align: center;
        }

        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(37, 99, 235, 0.3);
        }

        .btn-secondary {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            text-decoration: none;
            background: rgba(255, 255, 255, 0.08);
            color: var(--text-primary);
            border: 1px solid var(--glass-border);
            padding: 14px;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s ease;
            flex: 1;
            text-align: center;
        }

        .btn-secondary:hover {
            background: rgba(255, 255, 255, 0.15);
            border-color: rgba(255, 255, 255, 0.2);
        }

        /* Alerts */
        .alert-container {
            margin-bottom: 25px;
            animation: fadeIn 0.5s ease-out;
        }

        .alert {
            padding: 14px 20px;
            border-radius: 10px;
            font-weight: 500;
            font-size: 14px;
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 10px;
            border: 1px solid transparent;
        }

        .alert.danger {
            background: rgba(239, 68, 68, 0.15);
            color: var(--danger);
            border-color: rgba(239, 68, 68, 0.2);
        }

        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
    </style>
</head>
<body>
    <div class="edit-container">
        <div class="edit-card">
            <div class="edit-header">
                <i class="fas fa-user-edit"></i>
                <h2>Edit Employee</h2>
            </div>

            <!-- Flash Messages -->
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    <div class="alert-container">
                        {% for category, message in messages %}
                            <div class="alert {{ category }}">
                                <i class="fas fa-exclamation-circle"></i>
                                {{ message }}
                            </div>
                        {% endfor %}
                    </div>
                {% endif %}
            {% endwith %}

            <form action="/update/{{ employee.id }}" method="POST">
                <div class="form-group">
                    <label>Employee ID</label>
                    <input
                        type="text"
                        value="{{ employee.id }}"
                        readonly
                        class="readonly-input">
                </div>

                <div class="form-group">
                    <label>Employee Name</label>
                    <input
                        type="text"
                        name="employee_name"
                        value="{{ employee.name }}"
                        required>
                </div>

                <div class="form-group">
                    <label>Department</label>
                    <input
                        type="text"
                        name="department"
                        value="{{ employee.department }}"
                        required>
                </div>

                <div class="form-group">
                    <label>Salary</label>
                    <input
                        type="number"
                        name="salary"
                        value="{{ employee.salary }}"
                        required>
                </div>

                <div class="button-group">
                    <button type="submit" class="btn-primary">
                        <i class="fas fa-save"></i> Save
                    </button>
                    <a href="/" class="btn-secondary">
                        <i class="fas fa-arrow-left"></i> Cancel
                    </a>
                </div>
            </form>
        </div>
    </div>
</body>
</html>
"""

# ==========================================
# Routes
# ==========================================

# -----------------------------
# Home Page
# -----------------------------
@app.route("/")
def index():
    employees = get_all_employees()
    stats = get_dashboard_stats()
    return render_template_string(INDEX_HTML, employees=employees, stats=stats)


# -----------------------------
# Add Employee
# -----------------------------
@app.route("/add", methods=["POST"])
def add_employee():
    employee_id = request.form["employee_id"].strip()
    employee_name = request.form["employee_name"].strip()
    department = request.form["department"].strip()
    salary = request.form["salary"].strip()

    # Validation
    if not employee_id or not employee_name or not department or not salary:
        flash("All fields are required.", "danger")
        return redirect(url_for("index"))

    try:
        float(salary)
    except:
        flash("Salary must be numeric.", "danger")
        return redirect(url_for("index"))

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check Duplicate Employee ID
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == employee_id:
            flash("Employee ID already exists.", "danger")
            wb.close()
            return redirect(url_for("index"))

    ws.append([
        int(employee_id),
        employee_name,
        department,
        float(salary)
    ])

    wb.save(EXCEL_FILE)
    wb.close()

    flash("Employee added successfully.", "success")
    return redirect(url_for("index"))


# -----------------------------
# Edit Employee Page
# -----------------------------
@app.route("/edit/<employee_id>")
def edit_employee(employee_id):
    employees = get_all_employees()
    employee = None

    for emp in employees:
        if str(emp["id"]) == str(employee_id):
            employee = emp
            break

    if employee is None:
        flash("Employee not found.", "danger")
        return redirect(url_for("index"))

    return render_template_string(EDIT_HTML, employee=employee)


# -----------------------------
# Update Employee
# -----------------------------
@app.route("/update/<employee_id>", methods=["POST"])
def update_employee(employee_id):
    name = request.form["employee_name"].strip()
    department = request.form["department"].strip()
    salary = request.form["salary"].strip()

    if not name or not department or not salary:
        flash("All fields are required.", "danger")
        return redirect(url_for("edit_employee", employee_id=employee_id))

    try:
        float(salary)
    except:
        flash("Salary must be numeric.", "danger")
        return redirect(url_for("edit_employee", employee_id=employee_id))

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    found = False

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(employee_id):
            row[1].value = name
            row[2].value = department
            row[3].value = float(salary)
            found = True
            break

    if found:
        wb.save(EXCEL_FILE)
        flash("Employee updated successfully.", "success")
    else:
        flash("Employee not found.", "danger")

    wb.close()
    return redirect(url_for("index"))


# -----------------------------
# Delete Employee
# -----------------------------
@app.route("/delete/<employee_id>")
def delete_employee(employee_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    row_to_delete = None

    for row_num in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_num, column=1).value) == str(employee_id):
            row_to_delete = row_num
            break

    if row_to_delete:
        ws.delete_rows(row_to_delete, 1)
        wb.save(EXCEL_FILE)
        flash("Employee deleted successfully.", "success")
    else:
        flash("Employee not found.", "danger")

    wb.close()
    return redirect(url_for("index"))


# -----------------------------
# Run App
# -----------------------------
if __name__ == "__main__":
    create_excel_file()
    app.run(debug=True)
